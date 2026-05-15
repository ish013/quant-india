"""
excel_writer.py — Write scores to Excel with RSI + Strong Buy columns
"""

import logging
import platform
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import config as cfg

log = logging.getLogger(__name__)

_THIN   = Side(style="thin", color="CCCCCC")
_BORDER = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)

_FILL = {
    "title":        PatternFill("solid", fgColor="1F3864"),
    "header":       PatternFill("solid", fgColor="2E75B6"),
    "macro":        PatternFill("solid", fgColor="404040"),
    "alt":          PatternFill("solid", fgColor="EBF3FB"),
    "BUY":          PatternFill("solid", fgColor="C6EFCE"),
    "SELL":         PatternFill("solid", fgColor="FFC7CE"),
    "HOLD":         PatternFill("solid", fgColor="FFEB9C"),
    "STRONG BUY":   PatternFill("solid", fgColor="00FF88"),
    "NA":           PatternFill("solid", fgColor="F2F2F2"),
    "legend_title": PatternFill("solid", fgColor="D6DCE4"),
    "rsi_low":      PatternFill("solid", fgColor="C6EFCE"),
    "rsi_high":     PatternFill("solid", fgColor="FFC7CE"),
}

_FONT = {
    "title":       Font(bold=True,   color="FFFFFF", size=14, name="Calibri"),
    "header":      Font(bold=True,   color="FFFFFF", size=10, name="Calibri"),
    "macro":       Font(bold=True,   color="FFFFFF", size=10, name="Calibri"),
    "data":        Font(size=10,     name="Calibri"),
    "BUY":         Font(bold=True,   color="375623", size=10, name="Calibri"),
    "SELL":        Font(bold=True,   color="9C0006", size=10, name="Calibri"),
    "HOLD":        Font(bold=True,   color="7D6608", size=10, name="Calibri"),
    "STRONG BUY":  Font(bold=True,   color="004225", size=10, name="Calibri"),
    "NA":          Font(italic=True, color="808080", size=10, name="Calibri"),
    "asset":       Font(bold=True,   size=10, name="Calibri"),
    "ts":          Font(italic=True, color="FFFFFF", size=9,  name="Calibri"),
    "legend":      Font(size=9, name="Calibri"),
}

_C = Alignment(horizontal="center", vertical="center")
_L = Alignment(horizontal="left",   vertical="center")

# ── Columns (added RSI Daily + Strong Buy) ─────────────────────────────────
COLS = [
    ("A", "Asset",        13),
    ("B", "Price (₹)",    13),
    ("C", "52W High",     12),
    ("D", "52W Low",      12),
    ("E", "Price Score",  12),
    ("F", "P/E Ratio",    11),
    ("G", "PE Score",     11),
    ("H", "VIX",          10),
    ("I", "VIX Score",    11),
    ("J", "USD/INR (₹)",  12),
    ("K", "FX Score",     11),
    ("L", "RSI Daily",    11),   # ← NEW
    ("M", "Master Score", 13),
    ("N", "Signal",       14),   # ← STRONG BUY added
]

ASSET_ROWS = {
    "Nifty":    3,
    "Infosys":  4,
    "Reliance": 5,
    "ICICI":    6,
}

MACRO_ROW  = 8
LEGEND_ROW = 10
_cycle     = 0


def write_dashboard(scores: dict):
    global _cycle
    _cycle += 1
    _ensure_file()
    _write_openpyxl(scores)


def _write_openpyxl(scores: dict):
    path = Path(cfg.EXCEL_PATH)
    wb   = openpyxl.load_workbook(path)

    if cfg.SHEET_NAME in wb.sheetnames:
        ws = wb[cfg.SHEET_NAME]
    else:
        ws = wb.active
        ws.title = cfg.SHEET_NAME

    _init_sheet_openpyxl(ws)

    now = datetime.now().strftime("%d-%b-%Y  %H:%M:%S")
    _set_cell(ws, 1, 16, f"Updated: {now}   Cycle #{_cycle}",
              font=_FONT["ts"], align=_L)

    for name, row in ASSET_ROWS.items():
        s    = scores.get(name, {})
        rsi  = s.get("rsi")
        sbuy = s.get("strong_buy", False)
        sig  = "★ STRONG BUY" if sbuy else s.get("signal", "N/A")

        data_row = [
            name,
            _fmt(s.get("price"),        2, True),
            _fmt(s.get("high_52w"),     2, True),
            _fmt(s.get("low_52w"),      2, True),
            _fmt(s.get("price_score"),  3),
            _fmt(s.get("pe"),           2) or "N/A*",
            _fmt(s.get("pe_score"),     3) or "N/A*",
            _fmt(s.get("vix"),          2),
            _fmt(s.get("vix_score"),    3),
            _fmt(s.get("usdinr"),       4),
            _fmt(s.get("fx_score"),     3),
            _fmt(rsi,                   1) if rsi is not None else "N/A",  # RSI
            _fmt(s.get("master_score"), 3),
            sig,
        ]

        fill_alt = _FILL["alt"] if (row % 2 == 0) else None

        for col_idx, val in enumerate(data_row, start=1):
            is_sig  = (col_idx == 14)
            is_name = (col_idx == 1)
            is_rsi  = (col_idx == 12)

            if is_sig:
                sig_key = "STRONG BUY" if sbuy else s.get("signal", "N/A")
                fill = _FILL.get(sig_key, fill_alt)
                font = _FONT.get(sig_key, _FONT["data"])
            elif is_rsi and rsi is not None:
                fill = _FILL["rsi_low"]  if rsi < 30 else (
                       _FILL["rsi_high"] if rsi > 70 else fill_alt)
                font = Font(bold=True, color="375623", size=10, name="Calibri") if rsi < 30 else (
                       Font(bold=True, color="9C0006", size=10, name="Calibri") if rsi > 70 else _FONT["data"])
            elif is_name:
                fill = fill_alt
                font = _FONT["asset"]
            else:
                fill = fill_alt
                font = _FONT["data"]

            _set_cell(ws, row, col_idx, val,
                      font=font, fill=fill, align=_C, border=_BORDER)

    # Macro row
    macro = scores.get("macro", next(iter(scores.values()), {}))
    macro_vals = [
        "MACRO", "", "", "", "", "", "",
        _fmt(macro.get("vix"),       2),
        _fmt(macro.get("vix_score"), 3),
        _fmt(macro.get("usdinr"),    4),
        _fmt(macro.get("fx_score"),  3),
        "", "", "",
    ]
    for col_idx, val in enumerate(macro_vals, start=1):
        _set_cell(ws, MACRO_ROW, col_idx, val,
                  font=_FONT["macro"], fill=_FILL["macro"],
                  align=_C, border=_BORDER)

    try:
        wb.save(path)
        log.info("Excel saved → %s (cycle #%d)", path, _cycle)
        print(f"  ✅ Excel saved → {path.resolve()}")
    except PermissionError:
        log.warning("⚠ Excel file is open — close it to allow updates.")
        print("  ⚠ Excel is open — close it and it updates next cycle.")


def _init_sheet_openpyxl(ws):
    _set_cell(ws, 1, 1,
              "📊  QUANTITATIVE ANALYZER  —  Nifty | Infosys | Reliance | ICICI",
              font=_FONT["title"], fill=_FILL["title"], align=_L)
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:N1")

    for i, (_, hdr, width) in enumerate(COLS, start=1):
        _set_cell(ws, 2, i, hdr,
                  font=_FONT["header"], fill=_FILL["header"],
                  align=_C, border=_BORDER)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[2].height = 20

    for name, row in ASSET_ROWS.items():
        _set_cell(ws, row, 1, name, font=_FONT["asset"], align=_C, border=_BORDER)

    _set_cell(ws, MACRO_ROW, 1, "MACRO DATA",
              font=_FONT["macro"], fill=_FILL["macro"], align=_C, border=_BORDER)

    _set_cell(ws, LEGEND_ROW, 1, "SCORE LEGEND",
              font=Font(bold=True, name="Calibri", size=10),
              fill=_FILL["legend_title"], align=_L)

    legends = [
        ("★ STRONG BUY — RSI < 30  AND  Score < 0.35",   "STRONG BUY"),
        ("🟢 BUY  — Score < 0.35   (cheap + opportunity)", "BUY"),
        ("🟡 HOLD — Score 0.35 – 0.65",                    "HOLD"),
        ("🔴 SELL — Score > 0.65   (expensive)",           "SELL"),
        ("RSI < 30 = Oversold (green)  |  RSI > 70 = Overbought (red)", "NA"),
        ("N/A* — P/E Score requires paid API",             "NA"),
    ]
    for i, (text, key) in enumerate(legends):
        _set_cell(ws, LEGEND_ROW + 1 + i, 1, text,
                  font=_FONT["legend"], fill=_FILL[key], align=_L)
        ws.column_dimensions["A"].width = max(
            ws.column_dimensions["A"].width, len(text) * 0.85
        )

    ws.freeze_panes = "A3"


def _ensure_file():
    path = Path(cfg.EXCEL_PATH)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cfg.SHEET_NAME
    wb.save(path)


def _set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    return c


def _fmt(val, decimals=2, thousands=False):
    if val is None: return ""
    try:
        v = float(val)
        return f"{v:,.{decimals}f}" if thousands else f"{v:.{decimals}f}"
    except (TypeError, ValueError):
        return str(val)